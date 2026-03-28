# Excel版学生管理系统
# (28号新增)
# 给系统加登录密码功能：打开程序必须先输入密码，错三次直接退出，更安全
# 系统全面加固：禁止添加同名学生，菜单输字母不会崩溃，代码更规范、更专业



# 导入库
from openpyxl import Workbook, load_workbook
import os

# 全局变量： 存所有学生
students = []

# ----------------------------
# 1. 初始化 Excel 文件(没有就创建)
# ----------------------------

def init_excel():
    # 如果文件不存在，创建一个带表头的 Excel
    if not os.path.exists("学生表.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["姓名","年龄","成绩"]) # 表头
        wb.save("学生表.xlsx")
        print("首次运行已创建 学生表.xlsx")

# ----------------------------
# 2. 从 Excel 读取所有学生
# ----------------------------

def load_from_excel():
    global students
    wb = load_workbook("学生表.xlsx")
    ws = wb.active
    students.clear()
    for row in ws.iter_rows(min_row = 2,values_only=True):
        name , age , score = row
        stu = {
            "name" : name,
            "age" : age,
            "score" : score
        }
        students.append(stu)
    print("已从 Excel 加载学生数据")

# ----------------------------
# 3. 保存所有学生回 Excel
# ----------------------------

def save_to_excel():
    wb = Workbook()
    ws = wb.active
    # 再写表头
    ws.append(["姓名","年龄","成绩"])
    # 再写所有学生
    for stu in students:
        ws.append([stu["name"],stu["age"],stu["score"]])
    wb.save("学生表.xlsx")

# ----------------------------
# 判断是不是数字
# ----------------------------
def is_number(s):
    try:
        float(s)
        return True
    except:
        return False

# ----------------------------
# 判断姓名是否重复
# ----------------------------

def is_name_exist(name):
    # 从 students 中拿出元素，依次放到stu里面
    for stu in students:
        if stu["name"] == name:
            return True
        return False

# ----------------------------
# 4. 添加学生(带重名校验)
# ----------------------------

# def add_student():
#     name = input("姓名：")
#     age = input("年龄：")
#     score = input("成绩: ")
#     stu = {
#         "name" : name,
#         "age" : age,
#         "score" : score
#     }
#     students.append(stu)
#     save_to_excel()
#     print("添加成功！\n")

def add_student():
    print("----- 添加学生 -----")
    name = input("请输入姓名：").strip()
    if not name:
        print("姓名不能为空！")
        return
    if is_name_exist(name):
        print("该学生已经存在，不能重复添加！")
        return
    age = input("请输入年龄：").strip()
    if not age.isdigit():
        print("年龄必须是数字！")
        return
    score = input("请输入成绩：").strip()
    if not is_number(score):
        print("成绩必须是数字！")
        return
    # 类型转换
    age = int(age)
    score = float(score)
    stu = {"name":name, "age":age, "score":score}
    students.append(stu)
    save_to_excel()
    print("添加成功！")


# ----------------------------
# 5. 查看所有学生
# ----------------------------

def show_all():
    if not students:
        print("暂无学生信息！\n")
        return
    print("------ 所有学生------")
    for i,stu in enumerate(students,1):
        print(f"{i},姓名：{stu['name']},年龄：{stu['name']},成绩{stu['name']}")
    print()

# ----------------------------
# 6. 查找学生
# ----------------------------

def find_student():
    name = input("输入要查找的姓名：")
    for stu in students:
        if stu["name"] == name:
            print(f"找到了，{stu["name"]},年龄：{stu["age"]},成绩：{stu["score"]}")
            print()
            return
    print("未找到该学生！\n")

# ----------------------------
# 7. 删除学生
# ----------------------------

def delete_student():
    name = input("输入要删除的姓名：").strip()
    for stu in students:
        if stu["name"] == name:
            students.remove(stu)
            save_to_excel()
            print(f"已删除{name}")
            return
    print("未找到该学生！\n")

# ----------------------------
# 8. 修改成绩(新，带异常处理)
# ----------------------------

# def modify_score():
#     name = input("输入要修改成绩的学生姓名：")
#     for stu in students:
#         if stu["name"] == name:
#             new_score = input("输入修改后的成绩：")
#             stu["score"] = new_score
#             save_to_excel()
#             print("修改成功！\n")
#             return
#     print("未找到该学生！\n")

def modify_score():
    name = input("输入要修改成绩的学生姓名：").strip()
    for stu in students:
        if stu["name"] == name:
            new_score = input("请输入新的成绩：").strip()
            if not is_number(new_score):
                print("成绩必须是数字！")
                return
            stu["score"] = float(new_score)
            save_to_excel()
            print("修改成功！")
            return
    print("未找到该学生！")

# ----------------------------
#  成绩统计功能
# ----------------------------

def score_statistics():
    if not students:
        print("暂无学生数据，无法统计！")
        return
    # 1. 提取所有成绩
    scores = [stu["score"] for stu in students]
    # 2. 计算统计值
    total = sum(scores)
    avg = total / len(students)
    highest = max(scores)
    lowest = min(scores)
    # 3. 按成绩降序排名
    ranked = sorted(students, key = lambda x: x["score"],reverse = True)
    # 4. 打印报表
    print("\n" + "="*30)
    print("学生成绩统计报表")
    print("="*30)
    print(f"总人数：{len(students)}")
    print(f"总分：{total:.2f}")
    print(f"平均分：{avg:.2f}")
    print(f"最高分：{highest}")
    print(f"最低分：{lowest}")
    print("\n----- 成绩排名 -----")
    for i,stu in enumerate(ranked,1):
        print(f"第{i}名：{stu['name']} - {stu['score']}分")
    print("="*30 + "\n")

# ----------------------------
#  密码登录
# ----------------------------
def login():
    # 正确密码
    right_pwd = "123456"
    # 最多输入3次
    max_times = 3
    # 已经输了几次
    count = 0
    # 循环让用户输密码
    while count < max_times:
        pwd = input("请输入密码：").strip()
        if pwd == right_pwd:
            print("登陆成功！欢迎使用学生管理系统！")
            return True
        else:
            # 错误次数+1
            count += 1
            print(f"密码错误！你还剩{max_times - count}次机会")
    # 3次都错
    print("3次密码错误，程序自动退出")
    return False



# ----------------------------
# 9. 程序入口(带登录)(菜单防崩溃)
# ----------------------------

if __name__ == '__main__':
    # 先登录
    if not login():
        # 登录失败直接退出
        exit()
    # 登录成功才往下走
    init_excel()
    load_from_excel()
    while True:
        print("------ 学生管理系统 ------")
        print("1. 添加学生")
        print("2. 查看所有学生")
        print("3. 查找学生")
        print("4. 删除学生")
        print("5. 修改成绩")
        print("6. 成绩统计与排名")
        print("0. 退出")
        choice = input("请输入序号：").strip()
        if not choice.isdigit():
            print("请输入数字选项！")
            continue
        # 转成整数
        choice = int(choice)
        if choice == 1:
            add_student()
        elif choice == 2:
            show_all()
        elif choice == 3:
            find_student()
        elif choice == 4:
            delete_student()
        elif choice == 5:
            modify_score()
        elif choice == 6:
            score_statistics()
        elif choice == 0:
            save_to_excel()
            print("已退出，数据已保存到 Excel")
            break
        else:
            print("输入错误，请重新输入\n")








